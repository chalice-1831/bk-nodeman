# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-节点管理(BlueKing-BK-NODEMAN) available.
Copyright (C) 2017-2022 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at https://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
import logging
from typing import Any, Dict

from openpyxl import Workbook

from apps.node_man import constants, models
from apps.node_man.handlers.cmdb import CmdbHandler
from apps.node_man.tools.excel import ExcelTools

logger = logging.getLogger("app")

MAIN_SHEET_NAME = "bk_nodeman_info"

ANALYZE_ERROR_MSG = "第{}行：{}不能为空或格式错误，请依照模板填写，不要修改模板"


class ExcelHandler:
    @classmethod
    def generate_excel_template(cls):

        # 整合数据转为下拉框所需列表
        all_install_channel = [constants.DEFAULT_INSTALL_CHANNEL_NAME] + list(
            models.InstallChannel.install_channel_id_name_map().values()
        )
        all_biz = [item["bk_biz_name"] for item in CmdbHandler().biz(param={"action": "agent_operate"})]
        all_cloud = [constants.DEFAULT_CLOUD_NAME] + [
            cloud.bk_cloud_name for cloud in models.Cloud.objects.all().only("bk_cloud_name")
        ]
        all_ap = [constants.AUTOMATIC_CHOICE] + [ap.name for ap in models.AccessPoint.objects.all().only("name")]
        all_os = list(constants.OsType)
        all_auth_type = [str(type) for type in constants.ExcelAuthType.get_member_value__alias_map().values()]
        all_addressing = [str(type) for type in constants.CmdbAddressingType.get_member_value__alias_map().values()]
        all_enable_compression = ["True", "False"]

        # 生成excel模板
        excel = Workbook()
        excel_sheet = excel.active
        excel_sheet.title = MAIN_SHEET_NAME

        excel_field: Dict[Any, str] = constants.ExcelField._get_member__alias_map()
        excel_optional = constants.ExcelOptionalType._get_member__alias_map()
        excel_field_optional = constants.ExcelField.get_excel_optional_map()
        excel_describe = constants.ExcelField.get_excel_describe_map()
        for col, key in enumerate(constants.ExcelField, start=1):
            title_row_cell = excel_sheet.cell(row=1, column=col, value=str(excel_field[key]))
            ExcelTools.set_font_style(title_row_cell, font_size=16, color="538DD5", bold=True)

            optional_row_cell = excel_sheet.cell(row=2, column=col, value=str(excel_field_optional[key]))
            if excel_field_optional[key] == excel_optional[constants.ExcelOptionalType.REQUIRED]:
                ExcelTools.set_font_style(optional_row_cell, font_size=12, color="C0504D")
            else:
                ExcelTools.set_font_style(optional_row_cell, font_size=12, color="E26B0A")

            describe_row_cell = excel_sheet.cell(row=3, column=col, value=str(excel_describe[key]))
            ExcelTools.set_font_style(describe_row_cell, font_size=12, color="000000")

            if key == constants.ExcelField.OS_TYPE:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_os)
            elif key == constants.ExcelField.INSTALL_CHANNEL:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_install_channel)
            elif key == constants.ExcelField.AUTH_TYPE:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_auth_type)
            elif key == constants.ExcelField.BIZ:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_biz)
            elif key == constants.ExcelField.CLOUD:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_cloud)
            elif key == constants.ExcelField.AP:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_ap)
            elif key == constants.ExcelField.ADDRESS_TYPE:
                ExcelTools.create_dropdown(excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_addressing)
            elif key == constants.ExcelField.DATA_COMPRESSION:
                ExcelTools.create_dropdown(
                    excel, 4, col, str(excel_field[key]), MAIN_SHEET_NAME, all_enable_compression
                )
            else:
                pass

        ExcelTools.fill_color(excel_sheet, 1, 3, 1, len(excel_field), "D9D9D9")
        # 调整首行高度 25 次行 35 描述行 175 宽度 35
        ExcelTools.adjust_row_height(excel_sheet, 1, 1, 30)
        ExcelTools.adjust_row_height(excel_sheet, 2, 2, 35)
        ExcelTools.adjust_row_height(excel_sheet, 3, 3, 175)
        ExcelTools.adjust_col_width(excel_sheet, 1, len(excel_field), 35)
        ExcelTools.set_alignment(excel_sheet, "center", "left")

        return excel
