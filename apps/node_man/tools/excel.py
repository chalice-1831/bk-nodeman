# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making 蓝鲸智云-节点管理(BlueKing-BK-NODEMAN) available.
Copyright (C) 2017-2022 THL A29 Limited, a Tencent company. All rights reserved.
Licensed under the MIT License (the "License"); you may not use this file except in compliance with the License.
You may obtain a copy of the License at https://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
specific language governing permissions and limitations under the License.
"""
from typing import List

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_DROP_DOWN_ROW = 1000


class ExcelTools:
    @classmethod
    def fill_color(
        cls,
        excel: Worksheet,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        color: str,
        fill_type: str = FILL_SOLID,
    ):
        fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                excel.cell(row=row, column=col).fill = fill

    @classmethod
    def create_dropdown(
        cls, excel: Workbook, start_row: int, col: int, src_sheet: str, dst_sheet: str, options: List[str]
    ):
        sheet = excel.create_sheet(title=src_sheet)
        main_sheet = excel[dst_sheet]
        for i, option in enumerate(options, start=1):
            sheet[f"A{i}"] = option

        dv = DataValidation(type="list", formula1=f"={src_sheet}!$A$1:$A${len(options)}", allow_blank=True)

        # 默认提供1000行数据下拉
        main_sheet.add_data_validation(dv)
        dv.add(f"{chr(64 + col)}{start_row}:{chr(64 + col)}{DEFAULT_DROP_DOWN_ROW}")

    @classmethod
    def adjust_row_height(cls, excel: Worksheet, start_row: int, end_row: int, height: float):
        for row in range(start_row, end_row + 1):
            excel.row_dimensions[row].height = height

    @classmethod
    def adjust_col_width(cls, excel: Worksheet, start_col: int, end_col: int, width: float):
        for col in range(start_col, end_col + 1):
            excel.column_dimensions[chr(64 + col)].width = width

    @classmethod
    def set_alignment(cls, excel: Worksheet, vertical: str, horizontal: str):
        alignment = Alignment(wrap_text=True, vertical=vertical, horizontal=horizontal)
        for row in excel.iter_rows():
            for cell in row:
                cell.alignment = alignment

    @classmethod
    def set_font_style(
        cls,
        cell: Cell,
        font_size: int,
        color: str = "000000",
        name: str = "SimSun",
        bold: bool = False,
        italic: bool = False,
        strike: bool = False,
    ):
        font = Font(size=font_size, color=color, name=name, bold=bold, italic=italic, strike=strike)
        cell.font = font
